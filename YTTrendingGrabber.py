import os
import requests
from time import sleep
import json
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from imageToBrightnessCalc import measureAverageBrightness



def main():
      currentDay = date.today()
      nextDay = date.today() + timedelta(days=1)


      foldername = str(date.today())
      if not os.path.isdir("thumbnails/" + foldername):
            os.mkdir("thumbnails/" + foldername)

            
            videoIDs = ["Video ID"]
            videoViews = ["Video Views"]
            videoBrightnesses = ["Video Average Brightness"]

            heading = [date.today()]


            wb = load_workbook("Youtube Trending Videos.xlsx")
            ws = wb.active

            youtubeRequest = requests.get("https://youtube.googleapis.com/youtube/v3/videos?part=snippet%2CcontentDetails%2Cstatistics&chart=mostPopular&regionCode=US&key=getYourOwnKey:D" )
            youtubeRequestJson = json.loads(youtubeRequest.text)

            for item in youtubeRequestJson['items']:

                  videoURL = item['snippet']['thumbnails']['default']['url']
                  youtubeThumbnailRequest = requests.get(videoURL)

                  videoIDs.append(item['id'])
                  videoViews.append(item['statistics']['viewCount'])

            
                  with open("thumbnails/"+foldername + "/" + item['id']+ ".jpg", 'wb') as youtubeThumbnailFile:
                        youtubeThumbnailFile.write(youtubeThumbnailRequest.content)

                  filePath = ".../thumbnails"
                  avgBright = measureAverageBrightness(filePath + "/"  +foldername + "/" + item['id']+ ".jpg")

                  videoBrightnesses.append(avgBright)

            
            ws.append(heading)
            ws.append(videoViews)
            ws.append(videoIDs)
            ws.append(videoBrightnesses)
            
            ws["A1"].value = ws["A1"].value + 1

            wb.save(filename = "Youtube Trending Videos.xlsx")



main()
                     
